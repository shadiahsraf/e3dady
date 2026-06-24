import io
import json
import zipfile
import secrets
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST, require_GET
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login as auth_login, logout as auth_logout, authenticate
from django.contrib.auth.models import User

import qrcode

from .models import Team, Participant, Activity, ScanLog, Session
from . import services


# ── Auth ──

def login_view(request):
    error = None
    if request.method == 'POST':
        u = request.POST.get('username', '').strip()
        p = request.POST.get('password', '').strip()

        if u == settings.LEADER_USERNAME and p == settings.LEADER_PASSWORD:
            user, _ = User.objects.get_or_create(
                username=settings.LEADER_USERNAME, defaults={'is_staff': False})
            user.set_password(settings.LEADER_PASSWORD)
            user.save()
            user = authenticate(request, username=u, password=p)
            if user:
                auth_login(request, user)
                return redirect('scanner')
        else:
            user = authenticate(request, username=u, password=p)
            if user:
                auth_login(request, user)
                return redirect('dashboard' if user.is_staff or user.is_superuser else 'scanner')

        error = 'بيانات الدخول غلط!'
    return render(request, 'core/login.html', {'error': error})


def logout_view(request):
    auth_logout(request)
    return redirect('login')


# ── Pages ──

def home(request):
    return redirect('leaderboard')


def scanner(request):
    activities = Activity.objects.all()
    session = services.get_active_session()
    active_id = session.activity.id if session else None
    if not active_id:
        active = activities.filter(is_active=True).first()
        active_id = active.id if active else None
    return render(request, 'core/scanner.html', {
        'activities': activities,
        'active_activity_id': active_id,
    })


def leaderboard(request):
    return render(request, 'core/leaderboard.html')


def team_view(request):
    return render(request, 'core/team_view.html')


def live_screen(request):
    return render(request, 'core/live_screen.html')


@login_required
def qr_codes_page(request):
    participants = Participant.objects.select_related('team').all()
    teams = Team.objects.all()
    return render(request, 'core/qr_codes.html', {
        'participants': participants,
        'teams': teams,
    })


@login_required
def dashboard(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return redirect('scanner')
    teams = Team.objects.all()
    activities = Activity.objects.all()
    participants = Participant.objects.select_related('team').all()
    session = services.get_active_session()
    sessions_history = Session.objects.select_related('activity').all()[:20]
    recent_logs = ScanLog.objects.select_related('participant', 'team').all()[:50]
    return render(request, 'core/dashboard.html', {
        'teams': teams,
        'activities': activities,
        'participants': participants,
        'active_session': session,
        'sessions_history': sessions_history,
        'recent_logs': recent_logs,
    })


# ── API: Lookup ──

@require_GET
def api_lookup(request):
    code = request.GET.get('code', '').strip()
    if not code:
        return JsonResponse({'ok': False, 'error': 'كود فاضي'}, status=400)

    result = services.lookup_code(code)
    if not result:
        return JsonResponse({'ok': False, 'error': 'الكود ده مش موجود'}, status=404)

    if result['type'] == 'individual':
        p = result['participant']
        return JsonResponse({
            'ok': True,
            'scan_type': 'individual',
            'participant': {
                'id': p.id,
                'name': p.name,
                'unique_code': p.unique_code,
                'total_points': p.total_points,
                'team': {
                    'id': p.team.id,
                    'name': p.team.name,
                    'name_ar': p.team.name_ar,
                    'color': p.team.color,
                    'primary': p.team.primary_hex,
                    'secondary': p.team.secondary_hex,
                    'flag': p.team.flag_emoji,
                },
            },
        })
    else:
        t = result['team']
        members = list(t.participants.values_list('name', flat=True))
        return JsonResponse({
            'ok': True,
            'scan_type': 'team',
            'team': {
                'id': t.id,
                'name': t.name,
                'name_ar': t.name_ar,
                'flag': t.flag_emoji,
                'primary': t.primary_hex,
                'total_points': t.total_points,
                'team_code': t.team_code,
                'members_count': t.participants.count(),
                'members': members,
            },
        })


# ── API: Record Scan ──

@csrf_exempt
@require_POST
def api_record_scan(request):
    code = request.POST.get('code', '').strip()
    try:
        points = int(request.POST.get('points', 1) or 1)
    except (ValueError, TypeError):
        points = 1
    location = request.POST.get('location', '')
    scanned_by = request.user.username if request.user.is_authenticated else 'leader'

    points_change = max(-999, min(999, points))

    result = services.lookup_code(code)
    if not result:
        return JsonResponse({'ok': False, 'error': 'الكود مش موجود'}, status=404)

    if result['type'] == 'individual':
        p = result['participant']
        log = services.record_individual_scan(p, points_change, location, scanned_by)
        return JsonResponse({
            'ok': True,
            'scan_type': 'individual',
            'points_change': log.points_change,
            'participant': {
                'id': p.id,
                'name': p.name,
                'total_points': p.total_points,
                'team_name_ar': p.team.name_ar,
                'team_total': p.team.total_points,
            },
            'message': f'اللاعب ده معاه {p.total_points} نقطة من بداية المؤتمر',
        })
    else:
        t = result['team']
        logs = services.record_team_scan(t, points_change, location, scanned_by)
        t.refresh_from_db()
        return JsonResponse({
            'ok': True,
            'scan_type': 'team',
            'points_change': points_change,
            'team': {
                'id': t.id,
                'name_ar': t.name_ar,
                'flag': t.flag_emoji,
                'total_points': t.total_points,
                'members_affected': len(logs),
            },
            'message': 'تم تسجيل الفريق بالكامل',
        })


# ── API: Data feeds ──

@require_GET
def api_leaderboard(request):
    teams = services.get_team_scoreboard()
    return JsonResponse({
        'teams': [
            {
                'id': t.id,
                'name': t.name,
                'name_ar': t.name_ar,
                'color': t.color,
                'primary': t.primary_hex,
                'secondary': t.secondary_hex,
                'flag': t.flag_emoji,
                'total': t.computed_total,
            } for t in teams
        ],
    })


@require_GET
def api_team_view(request):
    return JsonResponse(services.get_team_view_data())


@require_GET
def api_live_data(request):
    return JsonResponse(services.get_team_view_data())


# ── API: Session management ──

@login_required
@require_POST
def api_session_start(request):
    activity_id = request.POST.get('activity_id')
    try:
        activity = Activity.objects.get(pk=activity_id)
    except Activity.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'النشاط مش موجود'}, status=404)
    session = services.start_session(activity)
    return JsonResponse({'ok': True, 'session_id': session.id, 'activity': session.activity.label})


@login_required
@require_POST
def api_session_end(request):
    session = services.end_session()
    if not session:
        return JsonResponse({'ok': False, 'error': 'مفيش جلسة نشطة'}, status=400)
    return JsonResponse({'ok': True, 'session_id': session.id})


@login_required
@require_POST
def api_session_reset(request):
    session = services.reset_session_checkmarks()
    if not session:
        return JsonResponse({'ok': False, 'error': 'مفيش جلسة نشطة'}, status=400)
    return JsonResponse({'ok': True, 'session_id': session.id})


# ── API: Admin ──

@login_required
@require_POST
def api_add_participant(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'ok': False, 'error': 'مش مسموح'}, status=403)
    name = request.POST.get('name', '').strip()
    team_id = request.POST.get('team_id')
    if not name or not team_id:
        return JsonResponse({'ok': False, 'error': 'الاسم والفريق مطلوبين'}, status=400)
    try:
        team = Team.objects.get(pk=team_id)
    except Team.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'الفريق مش موجود'}, status=404)
    code = secrets.token_urlsafe(6)[:8].upper().replace('_', 'A').replace('-', 'B')
    p = Participant.objects.create(name=name, unique_code=code, team=team)
    return JsonResponse({
        'ok': True,
        'participant': {'id': p.id, 'name': p.name, 'unique_code': p.unique_code, 'team_name_ar': team.name_ar}
    })


@login_required
@require_POST
def api_delete_participant(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'ok': False, 'error': 'مش مسموح'}, status=403)
    pid = request.POST.get('participant_id')
    try:
        Participant.objects.get(pk=pid).delete()
    except Participant.DoesNotExist:
        return JsonResponse({'ok': False, 'error': 'مش موجود'}, status=404)
    return JsonResponse({'ok': True})


@login_required
@require_GET
def api_scan_logs(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({'ok': False}, status=403)
    participant_id = request.GET.get('participant_id')
    team_id = request.GET.get('team_id')
    location = request.GET.get('location')
    logs = services.get_scan_logs(participant_id, team_id, location)
    return JsonResponse({
        'ok': True,
        'logs': [
            {
                'id': l.id,
                'participant': l.participant.name if l.participant else None,
                'team': l.team.name_ar if l.team else None,
                'points_change': l.points_change,
                'session_type': l.get_session_type_display(),
                'location': l.get_location_display(),
                'scanned_by': l.scanned_by,
                'time': l.created_at.strftime('%H:%M:%S'),
            } for l in logs
        ],
    })


# ── QR codes ──

def _make_qr_png(data):
    img = qrcode.make(data, box_size=8, border=2)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return buf.getvalue()


def _scanner_url(request, code):
    scheme = 'https' if request.is_secure() else 'http'
    host = request.get_host()
    return f'{scheme}://{host}/scanner/?code={code}'


def qr_image(request, code):
    # Could be participant or team code
    try:
        Participant.objects.get(unique_code=code)
    except Participant.DoesNotExist:
        try:
            Team.objects.get(team_code=code)
        except Team.DoesNotExist:
            from django.http import Http404
            raise Http404
    url = _scanner_url(request, code)
    png = _make_qr_png(url)
    return HttpResponse(png, content_type='image/png')


def qr_team_image(request, team_code):
    team = get_object_or_404(Team, team_code=team_code)
    url = _scanner_url(request, team.team_code)
    png = _make_qr_png(url)
    return HttpResponse(png, content_type='image/png')


def qr_download_all(request):
    base_scheme = 'https' if request.is_secure() else 'http'
    base_host = request.get_host()
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for p in Participant.objects.select_related('team').all():
            url = f'{base_scheme}://{base_host}/scanner/?code={p.unique_code}'
            png = _make_qr_png(url)
            safe_name = p.name.replace('/', '_').replace(' ', '_')
            zf.writestr(f"{p.team.name}/{safe_name}_{p.unique_code}.png", png)
        for t in Team.objects.all():
            if t.team_code:
                url = f'{base_scheme}://{base_host}/scanner/?code={t.team_code}'
                png = _make_qr_png(url)
                zf.writestr(f"teams/TEAM_{t.name}_{t.team_code}.png", png)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/zip')
    resp['Content-Disposition'] = 'attachment; filename="qr_codes_all.zip"'
    return resp


# ── Excel export ──

@login_required
def export_points_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()

    # Sheet 1: All players summary
    ws = wb.active
    ws.title = 'نقاط اللاعبين'
    ws.sheet_view.rightToLeft = True

    header_font = Font(name='Cairo', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1D4ED8', end_color='1D4ED8', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = ['#', 'الاسم', 'الفريق', 'الكود', 'إجمالي النقاط']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    participants = Participant.objects.select_related('team').order_by('team__name', 'name')
    for i, p in enumerate(participants, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=p.name).border = thin_border
        ws.cell(row=row, column=3, value=p.team.name_ar).border = thin_border
        ws.cell(row=row, column=4, value=p.unique_code).border = thin_border
        pts_cell = ws.cell(row=row, column=5, value=p.total_points)
        pts_cell.border = thin_border
        pts_cell.alignment = Alignment(horizontal='center')
        pts_cell.font = Font(bold=True, size=12)

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16

    # Sheet 2: Scan log history
    ws2 = wb.create_sheet('سجل المسح')
    ws2.sheet_view.rightToLeft = True
    log_headers = ['الوقت', 'اللاعب', 'الفريق', 'النقاط', 'النوع', 'المكان', 'بواسطة']
    for col_idx, h in enumerate(log_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color='7C3AED', end_color='7C3AED', fill_type='solid')
        cell.alignment = header_align
        cell.border = thin_border

    logs = ScanLog.objects.select_related('participant', 'team').order_by('-created_at')[:500]
    for i, log in enumerate(logs, 1):
        row = i + 1
        ws2.cell(row=row, column=1, value=log.created_at.strftime('%Y-%m-%d %H:%M:%S')).border = thin_border
        ws2.cell(row=row, column=2, value=log.participant.name if log.participant else '—').border = thin_border
        ws2.cell(row=row, column=3, value=log.team.name_ar if log.team else '—').border = thin_border
        pts_cell = ws2.cell(row=row, column=4, value=log.points_change)
        pts_cell.border = thin_border
        pts_cell.font = Font(color='22C55E' if log.points_change > 0 else 'EF4444', bold=True)
        ws2.cell(row=row, column=5, value=log.get_session_type_display()).border = thin_border
        ws2.cell(row=row, column=6, value=log.get_location_display()).border = thin_border
        ws2.cell(row=row, column=7, value=log.scanned_by or '').border = thin_border

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 22
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 12
    ws2.column_dimensions['G'].width = 12

    # Sheet 3: Team summary
    ws3 = wb.create_sheet('ملخص الفرق')
    ws3.sheet_view.rightToLeft = True
    team_headers = ['الفريق', 'عدد اللاعبين', 'إجمالي النقاط']
    for col_idx, h in enumerate(team_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
        cell.alignment = header_align
        cell.border = thin_border

    for i, t in enumerate(Team.objects.all().order_by('-total_points'), 1):
        row = i + 1
        ws3.cell(row=row, column=1, value=t.name_ar).border = thin_border
        ws3.cell(row=row, column=2, value=t.participants.count()).border = thin_border
        pts_cell = ws3.cell(row=row, column=3, value=t.total_points)
        pts_cell.border = thin_border
        pts_cell.font = Font(bold=True, size=14)
        pts_cell.alignment = Alignment(horizontal='center')

    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="e3dady_cup_points.xlsx"'
    return resp
